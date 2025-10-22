from pathlib import Path
import yaml
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

TEMPLATE_DIR = "include"
TEMPLATE_FILE = "template.yml"

DAGS_WORKBOOK = "include/xlsx/dags.xlsx"
OUT = "dags/dynamic_dags.yml"

def generate_dags_from_template():
    # setup Jinja2
    env = Environment(loader=FileSystemLoader(TEMPLATE_DIR), autoescape=True)
    template = env.get_template(TEMPLATE_FILE)

    # generate dag per row
    workbook = load_workbook(DAGS_WORKBOOK)
    sheet = workbook.active

    all_dags = {}
    for row in sheet.iter_rows(min_row=3):
        dag_id = row[0].value
        command_1 = row[1].value
        command_2 = row[2].value

        if not dag_id:
            continue

        template_variables = {"dag_id": dag_id, "command_1": command_1, "command_2": command_2}
        dag_config = yaml.safe_load(template.render(template_variables))
        all_dags.update(dag_config)

    # write to file
    output_path = Path(OUT)
    with open(output_path, "w") as f:
        yaml.dump(all_dags, f, sort_keys=False)

if __name__ == "__main__":
    generate_dags_from_template()
